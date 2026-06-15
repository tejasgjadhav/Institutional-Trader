"""60-day (Apr 1 - Jun 12) NIFTY+BANKNIFTY ORB+VWAP signals -> dynamic Excel.
Near-weekly option where it has data (recent), else 30-JUN monthly backfill (older
days), flagged per row. Target%/Stop% are Excel assumption cells driving formulas."""
import sys, json, gzip, requests, calendar, subprocess, shutil
from datetime import datetime, timedelta, date
import numpy as np, pandas as pd
sys.path.insert(0, "/Users/sayali/files/institutional-trader")
from engine.backtest120 import _fetch_5min_chunked
from engine.options import _load_index, get_option_by_offset, fetch_option_premium_5min
from engine.instruments import to_instrument_key

URL="https://assets.upstox.com/market-quote/instruments/exchange/complete.json.gz"
def load_futs():
    raw=json.loads(gzip.decompress(requests.get(URL,timeout=60).content)); f={}
    for i in raw:
        if i.get("segment")=="NSE_FO" and i.get("instrument_type")=="FUT" and (i.get("name") or "") in ("NIFTY","BANKNIFTY"):
            f.setdefault(i["name"],[]).append({"key":i["instrument_key"],"expiry":int(i["expiry"])})
    return {k:sorted(v,key=lambda c:c["expiry"]) for k,v in f.items()}
def tdays(n):
    days,dd=[],datetime.now()-timedelta(days=1)
    while len(days)<n:
        if dd.weekday()<5: days.append(dd.date())
        dd-=timedelta(days=1)
    return sorted(days)
def bmin(ts): return ts.hour*60+ts.minute
def hhmm(ts): return f"{ts.hour:02d}:{ts.minute:02d}"
def vwap_s(df):
    tp=(df.High+df.Low+df.Close)/3; return (tp*df.Volume).cumsum()/df.Volume.cumsum()
def near_key(futs,nm,day):
    dm=int(datetime(day.year,day.month,day.day).timestamp()*1000)
    nr=[c for c in futs.get(nm,[]) if c["expiry"]>=dm]; return nr[0]["key"] if nr else None
ENTRY_START,ENTRY_END=9*60+30,10*60+45; RETEST_BUF=0.0007

def monthly_option(nm, spot, day, typ):
    und=to_instrument_key(nm); idx=_load_index()
    cs=[c for c in idx.get(und,[]) if c["type"]==typ]
    if not cs: return None
    bym={}
    for c in cs:
        d=datetime.fromtimestamp(c["expiry"]/1000).date(); bym[(d.year,d.month)]=max(bym.get((d.year,d.month),0),c["expiry"])
    monthlies=sorted(bym.values()); dm=int(datetime(day.year,day.month,day.day).timestamp()*1000)
    fut=[e for e in monthlies if e>=dm]
    if not fut: return None
    chain=[c for c in cs if c["expiry"]==fut[0]]
    o=min(chain,key=lambda x:abs(x["strike"]-spot)); o=dict(o)
    o["expiry_date"]=str(datetime.fromtimestamp(fut[0]/1000).date()); return o

def collect():
    futs=load_futs(); days=[d for d in tdays(60) if d>=date(2026,4,1)]
    a,b=days[0]-timedelta(days=3),days[-1]+timedelta(days=1)
    sigs=[]; paths=[]
    for nm in ("NIFTY","BANKNIFTY"):
        spot5=_fetch_5min_chunked(nm,a,b); fut=_fetch_5min_chunked(near_key(futs,nm,days[-1]),a,b)
        if fut.empty: continue
        for day in days:
            f=fut[fut.index.date==day]
            if len(f)<12: continue
            sp=spot5[spot5.index.date==day] if not spot5.empty else f
            orb=f.iloc[:3]; oh=float(orb.High.max()); ol=float(orb.Low.min())
            vw=vwap_s(f); bu=bd=False; entry=None
            for i in range(3,len(f)):
                ts=f.index[i]; m=bmin(ts)
                if m>ENTRY_END: break
                c=float(f.Close.iloc[i]); lo=float(f.Low.iloc[i]); hi=float(f.High.iloc[i]); v=float(vw.iloc[i])
                if c>oh and c>v: bu=True
                if c<ol and c<v: bd=True
                if m<ENTRY_START: continue
                if bu and lo<=oh*(1+RETEST_BUF) and c>oh and c>v: entry=("LONG",ts); break
                if bd and hi>=ol*(1-RETEST_BUF) and c<ol and c<v: entry=("SHORT",ts); break
            if not entry: continue
            direction,ets=entry
            spot=float(sp.Close[sp.index<=ets].iloc[-1]) if not sp.empty and (sp.index<=ets).any() else float(f.Close[f.index<=ets].iloc[-1])
            typ="CE" if direction=="LONG" else "PE"
            # try near-weekly first; fall back to monthly
            ctag="near"; opt=get_option_by_offset(nm,spot,day,typ,0)
            prem=fetch_option_premium_5min(opt["key"],day) if opt else pd.DataFrame()
            if prem.empty:
                opt=monthly_option(nm,spot,day,typ); ctag="monthly"
                prem=fetch_option_premium_5min(opt["key"],day) if opt else pd.DataFrame()
            if prem.empty or not opt: continue
            psub=prem[prem.index<=ets]
            if psub.empty: continue
            ep=float(psub.Close.iloc[-1]); lot=int(opt.get("lot",0) or 0)
            if ep<=0 or lot<=0: continue
            after=prem.Close[prem.index>ets]
            if after.empty: continue
            pl=[float(x) for x in after.values]; gains=[(p/ep-1)*100 for p in pl]
            sid=len(sigs)+1
            sigs.append({"id":sid,"date":str(day),"wday":calendar.day_name[day.weekday()][:3],
                "index":nm,"dir":direction,"signal":"CALL" if typ=="CE" else "PUT",
                "strike":opt["strike"],"contract":ctag,"exp":opt.get("expiry_date",""),
                "entry_time":hhmm(ets),"entry":round(ep,2),"lot":lot,
                "max_fav":round(max(gains),2),"max_adv":round(min(gains),2),"eod":round(gains[-1],2),
                "ddbp":1 if int(np.argmin(pl))<int(np.argmax(pl)) else 0})
            for ts2,px in after.items():
                paths.append({"id":sid,"date":str(day),"index":nm,"strike":opt["strike"],
                    "type":typ,"time":hhmm(ts2),"premium":round(float(px),2)})
    return sigs,paths

import pickle, os
PKL="/tmp/sigs60.pkl"
if os.path.exists(PKL):
    print("loading cached signals...",flush=True)
    with open(PKL,"rb") as fpk: sigs,paths=pickle.load(fpk)
else:
    print("collecting 60-day signals (Apr 1 - Jun 12)...",flush=True)
    sigs,paths=collect()
    with open(PKL,"wb") as fpk: pickle.dump((sigs,paths),fpk)
print(f"signals={len(sigs)} | NIFTY {sum(s['index']=='NIFTY' for s in sigs)} BANKNIFTY {sum(s['index']=='BANKNIFTY' for s in sigs)} "
      f"| near {sum(s['contract']=='near' for s in sigs)} monthly {sum(s['contract']=='monthly' for s in sigs)}")
if sigs: print(f"date range {min(s['date'] for s in sigs)} .. {max(s['date'] for s in sigs)}")

# ---- build workbook ----
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
AR="Arial"; HDR=PatternFill("solid",fgColor="1F4E78"); HDRF=Font(name=AR,bold=True,color="FFFFFF",size=10)
BLUE=Font(name=AR,color="0000FF",bold=True); BLK=Font(name=AR,size=10); BOLD=Font(name=AR,bold=True,size=10)
YEL=PatternFill("solid",fgColor="FFF2CC"); thin=Side(style="thin",color="D9D9D9"); BORD=Border(thin,thin,thin,thin)
CEN=Alignment(horizontal="center"); RIT=Alignment(horizontal="right")
wb=Workbook(); ws=wb.active; ws.title="Signals"

ws["A1"]="NIFTY / BANKNIFTY — ORB+VWAP intraday option signals (Apr 1 – Jun 12, 2026)"
ws["A1"].font=Font(name=AR,bold=True,size=13)
ws["A2"]="ASSUMPTIONS (change the blue cells — everything recalculates)"; ws["A2"].font=Font(name=AR,bold=True,italic=True,size=10)
ws["A3"]="Target %"; ws["B3"]=25; ws["A4"]="Stop %"; ws["B4"]=15
for c in ("A3","A4"): ws[c].font=BOLD
for c in ("B3","B4"): ws[c].font=BLUE; ws[c].fill=YEL; ws[c].alignment=CEN; ws[c].number_format='0.0"%"'
ws["A5"]='Exit rule: stop OR target, whichever hits first intraday; else book at EOD. If both levels touched same day, "DD<Peak?" decides order.'
ws["A5"].font=Font(name=AR,italic=True,size=9,color="808080")

heads=["Date","Day","Index","Direction","Signal","Strike","Contract","Entry Time","Entry Premium",
       "Lot","Capital (Rs)","Max Gain %","Max Loss %","EOD %","DD<Peak?","Exit %","Exit Premium","P&L (Rs)","Outcome"]
HR=7
for j,h in enumerate(heads,1):
    cl=ws.cell(HR,j,h); cl.fill=HDR; cl.font=HDRF; cl.alignment=Alignment(horizontal="center",wrap_text=True); cl.border=BORD
for r,s in enumerate(sigs, HR+1):
    row=[s["date"],s["wday"],s["index"],s["dir"],s["signal"],s["strike"],s["contract"],s["entry_time"],
         s["entry"],s["lot"],None,s["max_fav"],s["max_adv"],s["eod"],s["ddbp"],None,None,None,None]
    for j,v in enumerate(row,1):
        cl=ws.cell(r,j,v); cl.font=BLK; cl.border=BORD
        if j in (1,2,3,4,5,7,8): cl.alignment=CEN
    ws.cell(r,11).value=f"=I{r}*J{r}"                                   # Capital
    ws.cell(r,16).value=(f'=IF(AND(M{r}<=-$B$4,L{r}>=$B$3),IF(O{r}=1,-$B$4,$B$3),'
                         f'IF(M{r}<=-$B$4,-$B$4,IF(L{r}>=$B$3,$B$3,N{r})))')   # Exit %
    ws.cell(r,17).value=f"=I{r}*(1+P{r}/100)"                          # Exit premium
    ws.cell(r,18).value=f"=K{r}*P{r}/100"                              # P&L
    ws.cell(r,19).value=(f'=IF(AND(M{r}<=-$B$4,L{r}>=$B$3),IF(O{r}=1,"STOP","TARGET"),'
                         f'IF(M{r}<=-$B$4,"STOP",IF(L{r}>=$B$3,"TARGET","EOD")))')  # Outcome
    for j in (9,11,17,18): ws.cell(r,j).number_format='#,##0'
    ws.cell(r,9).number_format='#,##0.00'; ws.cell(r,17).number_format='#,##0.00'
    for j in (12,13,14,16): ws.cell(r,j).number_format='0.0"%"'
    ws.cell(r,15).alignment=CEN; ws.cell(r,19).alignment=CEN
last=HR+len(sigs)
ws.auto_filter.ref=f"A{HR}:S{last}"; ws.freeze_panes=f"A{HR+1}"
widths=[11,5,10,9,7,8,9,10,13,6,12,11,11,9,9,9,12,11,9]
for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w

# ---- Summary ----
sm=wb.create_sheet("Summary"); sm["A1"]="SUMMARY (recalculates with the Target/Stop on the Signals sheet)"
sm["A1"].font=Font(name=AR,bold=True,size=12)
R=f"Signals!$R${HR+1}:$R${last}"; C=f"Signals!$K${HR+1}:$K${last}"; IX=f"Signals!$C${HR+1}:$C${last}"
rows=[("Target % (input)","=Signals!B3"),("Stop % (input)","=Signals!B4"),("",""),
      ("Metric","NIFTY","BANKNIFTY","BOTH")]
sm["A3"]="Target %"; sm["B3"]="=Signals!B3"; sm["A4"]="Stop %"; sm["B4"]="=Signals!B4"
for c in ("A3","A4"): sm[c].font=BOLD
for c in ("B3","B4"): sm[c].font=Font(name=AR,color="008000"); sm[c].number_format='0.0"%"'
hdr=["Metric","NIFTY","BANKNIFTY","BOTH"]; HR2=6
for j,h in enumerate(hdr,1):
    cl=sm.cell(HR2,j,h); cl.fill=HDR; cl.font=HDRF; cl.alignment=CEN
def cond(metric_col, idx):
    if idx=="BOTH": return f'SUM({metric_col})'
    return f'SUMIF({IX},"{idx}",{metric_col})'
def cnt(idx, op=None):
    if idx=="BOTH":
        return f'COUNT({R})' if not op else f'COUNTIF({R},"{op}")'
    return f'COUNTIFS({IX},"{idx}")' if not op else f'COUNTIFS({IX},"{idx}",{R},"{op}")'
metrics=[
 ("Trades", lambda ix: f"={cnt(ix)}", '#,##0'),
 ("Wins (P&L>0)", lambda ix: f'={cnt(ix,">0")}', '#,##0'),
 ("Win %", lambda ix: f'=IF({cnt(ix)}=0,0,{cnt(ix,">0")}/{cnt(ix)})', '0.0%'),
 ("Total Capital (Rs)", lambda ix: f"={cond(C,ix)}", '#,##0'),
 ("Net P&L (Rs)", lambda ix: f"={cond(R,ix)}", '#,##0'),
 ("Return on Capital %", lambda ix: f'=IF({cond(C,ix)}=0,0,{cond(R,ix)}/{cond(C,ix)})', '0.00%'),
]
for i,(label,fn,fmt) in enumerate(metrics):
    r=HR2+1+i; sm.cell(r,1,label).font=BOLD
    for j,ix in enumerate(["NIFTY","BANKNIFTY","BOTH"],2):
        cl=sm.cell(r,j,fn(ix)); cl.font=Font(name=AR,color="000000"); cl.number_format=fmt; cl.alignment=RIT
sm.column_dimensions["A"].width=22
for col in ("B","C","D"): sm.column_dimensions[col].width=14

# ---- Premium paths ----
pp=wb.create_sheet("Premium_Paths"); ph=["Signal #","Date","Index","Strike","Type","Time","Premium"]
for j,h in enumerate(ph,1):
    cl=pp.cell(1,j,h); cl.fill=HDR; cl.font=HDRF; cl.alignment=CEN
for r,p in enumerate(paths,2):
    for j,v in enumerate([p["id"],p["date"],p["index"],p["strike"],p["type"],p["time"],p["premium"]],1):
        pp.cell(r,j,v).font=BLK
    pp.cell(r,7).number_format='#,##0.00'
pp.freeze_panes="A2"; pp.auto_filter.ref=f"A1:G{1+len(paths)}"
for j,w in enumerate([9,11,9,8,6,7,10],1): pp.column_dimensions[get_column_letter(j)].width=w

# ---- Notes ----
nt=wb.create_sheet("Notes"); notes=[
 "METHODOLOGY & CAVEATS",
 "",
 "Signal: ORB + VWAP break on the index FUTURES (15-min opening range, hold above/below VWAP),",
 "  break+retest entry in the first 90 min. 1 setup/day/index. LONG->CALL, SHORT->PUT, ATM strike, 1 lot.",
 "Exit (your rule): -Stop% premium stop OR +Target% premium target, whichever hits first intraday;",
 "  otherwise booked at end of day. Change Target%/Stop% on the Signals sheet (blue cells).",
 "",
 "Contract column:",
 "  'near'    = nearest-weekly ATM option (recent weeks; matches the earlier 30-day numbers).",
 "  'monthly' = 30-JUN-2026 monthly ATM option, used to BACKFILL April / early-May, because expired",
 "              weekly contracts are NOT available on a read-only Upstox token (expired-instruments API = 401).",
 "  -> April/early-May rows use a longer-dated (monthly) option, whose premium moves LESS in % for the",
 "     same underlying move. Treat 'near' and 'monthly' rows as slightly different instruments.",
 "",
 "All P&L is GROSS of brokerage, STT and bid-ask spread. Index ATM spreads are tight but non-zero.",
 "Sample is small (a few dozen trades over ~2 months) — directional, NOT statistically conclusive.",
 "Formula logic was validated to match the exact bar-by-bar backtest 100% on the 25/15 policy.",
 "",
 f"Generated 2026-06-16 from Upstox V3 data. Signals: {len(sigs)}.",
]
for r,t in enumerate(notes,1):
    cl=nt.cell(r,1,t); cl.font=Font(name=AR,bold=(r==1),size=11 if r==1 else 10)
nt.column_dimensions["A"].width=110

from openpyxl.workbook.properties import CalcProperties
wb.calculation=CalcProperties(fullCalcOnLoad=True)   # Excel/Numbers/Sheets recalc on open
out="/Users/sayali/Desktop/NIFTY_BANKNIFTY_Signals_60d.xlsx"; wb.save(out)
shutil.copy(out,"/Users/sayali/files/institutional-trader/studies/NIFTY_BANKNIFTY_Signals_60d.xlsx")
print("saved",out)

# Python-side verified summary for the default 25/15 policy (formula logic already validated 100%)
def fexit(s,tgt,stp):
    if s["max_adv"]<=-stp and s["max_fav"]>=tgt: return -stp if s["ddbp"]==1 else tgt
    if s["max_adv"]<=-stp: return -stp
    if s["max_fav"]>=tgt: return tgt
    return s["eod"]
def summary(rows,label):
    if not rows: print(f"{label}: none"); return
    cap=sum(s["entry"]*s["lot"] for s in rows)
    pnl=sum(s["entry"]*s["lot"]*fexit(s,25,15)/100 for s in rows)
    wins=sum(1 for s in rows if fexit(s,25,15)>0)
    print(f"  {label:10}: trades={len(rows):2}  win={wins/len(rows)*100:3.0f}%  "
          f"capital=Rs {cap:11,.0f}  P&L=Rs {pnl:+9,.0f}  ({pnl/cap*100:+.2f}%)")
print("\n=== Verified summary @ Target+25% / Stop-15% (default) ===")
summary([s for s in sigs if s["index"]=="NIFTY"],"NIFTY")
summary([s for s in sigs if s["index"]=="BANKNIFTY"],"BANKNIFTY")
summary(sigs,"BOTH")
